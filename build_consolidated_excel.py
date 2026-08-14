import glob
import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import sys

sys.stdout.reconfigure(encoding='utf-8')

def process_and_build():
    csv_files = glob.glob(r"g:\ΓΕΦΥΡΕΣ\ERGO ZWHS\1411-ΠΡΟΜΗΘΕΙΕΣ - ΥΠΕΡΠΡΟΜΗΘΕΙΕΣ *.csv")
    
    all_dfs = []
    for cf in sorted(csv_files):
        fname = os.path.basename(cf)
        month_code = fname.split(" ")[-1].replace(".csv", "").replace("_", "/")
        df = pd.read_csv(cf, encoding="cp1253", sep=";")
        df["Μήνας Εκκαθάρισης"] = month_code
        all_dfs.append(df)
        
    merged_df = pd.concat(all_dfs, ignore_index=True)
    
    def clean_num(val):
        if pd.isna(val):
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        val_str = str(val).strip().replace(".", "").replace(",", ".")
        try:
            return float(val_str)
        except:
            return 0.0

    num_cols = ["Καθαρά ΒΚ", "Καθαρά ΣΚ", "Καθαρά Σύνολο", "Προμήθεια ΒΚ", "Προμήθεια ΣΚ", "Προμήθεια Σύνολο", "Φόρος"]
    for col in num_cols:
        merged_df[col] = merged_df[col].apply(clean_num)
        
    merged_df["Date_Obj"] = pd.to_datetime(merged_df["Εναρξη"], format="%d/%m/%Y", errors="coerce")
    merged_df = merged_df.sort_values(by=["Date_Obj", "Συμβόλαιο"]).reset_index(drop=True)
    
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="E8EEF5")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=11, color="000000")

    fill_navy = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_teal = PatternFill(start_color="005F73", end_color="005F73", fill_type="solid")
    fill_accent_gold = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    thin_side = Side(border_style="thin", color="D9D9D9")
    double_bottom_side = Side(border_style="double", color="1B365D")
    thick_bottom_side = Side(border_style="medium", color="1B365D")

    border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    border_header = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thick_bottom_side)
    border_total = Border(top=thin_side, bottom=double_bottom_side)

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")

    # SHEET 1: Master Ledger
    ws1 = wb.create_sheet(title="Εκκαθαρίσεις ανά Ημερομηνία")
    ws1.views.sheetView[0].showGridLines = True
    
    ws1.merge_cells("A1:R1")
    ws1["A1"] = "ERGO - ΣΥΓΚΕΝΤΡΩΤΙΚΕΣ ΕΚΚΑΘΑΡΙΣΕΙΣ ΠΡΟΜΗΘΕΙΩΝ YΓΕΙΑΣ & ΖΩΗΣ ΑΝΑ ΗΜΕΡΟΜΗΝΙΑ"
    ws1["A1"].font = font_title
    ws1["A1"].fill = fill_navy
    ws1["A1"].alignment = align_center
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:R2")
    ws1["A2"] = "Ενοποιημένο αρχείο παραστατικών εκκαθάρισης (Φεβρουάριος 2026 - Ιούλιος 2026) σε χρονολογική σειρά"
    ws1["A2"].font = font_subtitle
    ws1["A2"].fill = fill_navy
    ws1["A2"].alignment = align_center
    ws1.row_dimensions[2].height = 20

    headers1 = [
        "Α/Α", "Ημερομηνία", "Μήνας Εκκαθ.", "Βαθμίδα", "Αρ. Συμβολαίου", "Αρ. Απόδειξης",
        "Ονοματεπώνυμο Πελάτη", "Τρ. Πληρ.", "Διαν. Έτος", "Διάρκεια",
        "Καθαρά ΒΚ (€)", "Καθαρά ΣΚ (€)", "Καθαρά Σύνολο (€)",
        "Προμήθεια ΒΚ (€)", "Προμήθεια ΣΚ (€)", "Προμήθεια Σύνολο (€)", "Φόρος (€)", "Καθαρό Πληρωτέο (€)"
    ]
    
    ws1.row_dimensions[4].height = 28
    for c_idx, h in enumerate(headers1, 1):
        cell = ws1.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = border_header

    for r_idx, row in merged_df.iterrows():
        row_num = r_idx + 5
        ws1.row_dimensions[row_num].height = 20
        client_fullname = f"{row['Επώνυμο']} {row['Ονομα.1'] if pd.notna(row['Ονομα.1']) else ''}".strip()
        date_str = row['Date_Obj'].strftime("%d/%m/%Y") if pd.notna(row['Date_Obj']) else str(row['Εναρξη'])
        
        vals = [
            r_idx + 1,
            date_str,
            row['Μήνας Εκκαθάρισης'],
            row['Βαθμίδα'],
            row['Συμβόλαιο'],
            row['Απόδειξη'],
            client_fullname,
            row['Τρ.Πληρ.'],
            row['Διαν.Ετος'],
            row['Διάρκεια'],
            row['Καθαρά ΒΚ'],
            row['Καθαρά ΣΚ'],
            row['Καθαρά Σύνολο'],
            row['Προμήθεια ΒΚ'],
            row['Προμήθεια ΣΚ'],
            row['Προμήθεια Σύνολο'],
            row['Φόρος'],
            f"=P{row_num}-Q{row_num}"
        ]
        
        for c_idx, val in enumerate(vals, 1):
            cell = ws1.cell(row=row_num, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_idx in [11, 12, 13, 14, 15, 16, 17, 18]:
                cell.number_format = "#,##0.00 €"
                cell.alignment = align_right
            elif c_idx in [1, 2, 3, 4, 5, 6, 8, 9, 10]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    tot_r1 = len(merged_df) + 5
    ws1.row_dimensions[tot_r1].height = 24
    ws1.cell(row=tot_r1, column=7, value="ΣΥΝΟΛΑ:").font = font_bold
    ws1.cell(row=tot_r1, column=7).alignment = align_right

    for c_i, col_let in [(11, "K"), (12, "L"), (13, "M"), (14, "N"), (15, "O"), (16, "P"), (17, "Q"), (18, "R")]:
        cell = ws1.cell(row=tot_r1, column=c_i, value=f"=SUM({col_let}5:{col_let}{tot_r1-1})")
        cell.font = font_bold
        cell.fill = fill_accent_gold
        cell.border = border_total
        cell.number_format = "#,##0.00 €"
        cell.alignment = align_right

    # SHEET 2: Enhanced Monthly Summary Table with Percentages
    ws2 = wb.create_sheet(title="Σύνολα ανά Μήνα")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "ERGO - ΜΗΝΙΑΙΑ ΠΟΣΟΣΤΑ & ΠΟΣΑ ΣΥΝΕΡΓΑΤΗ & AGENCY (02/2026 - 07/2026)"
    ws2["A1"].font = font_title
    ws2["A1"].fill = fill_navy
    ws2["A1"].alignment = align_center
    ws2.row_dimensions[1].height = 36

    ws2.merge_cells("A2:H2")
    ws2["A2"] = "Ανακεφαλαίωση καθαρών ασφαλίστρων, ποσών & ποσοστών προμηθειών ανά μήνα"
    ws2["A2"].font = font_subtitle
    ws2["A2"].fill = fill_navy
    ws2["A2"].alignment = align_center
    ws2.row_dimensions[2].height = 20

    headers2 = [
        "Μήνας Εκκαθάρισης", "Καθαρά Ασφάλιστρα (€)",
        "Προμήθεια Συνεργάτη (€)", "Ποσοστό Συνεργάτη (%)",
        "Override Agency (€)", "Ποσοστό Agency (%)",
        "Σύνολο Αμοιβών (€)", "Συνολικό Ποσοστό (%)"
    ]
    ws2.row_dimensions[4].height = 28
    for c_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = border_header

    months_list = ["02/2026", "03/2026", "04/2026", "05/2026", "06/2026", "07/2026"]
    
    for r_idx, m_code in enumerate(months_list, 5):
        ws2.row_dimensions[r_idx].height = 22
        ws2.cell(row=r_idx, column=1, value=m_code).alignment = align_center
        ws2.cell(row=r_idx, column=1).font = font_bold
        ws2.cell(row=r_idx, column=1).border = border_cell
        
        p_cell = ws2.cell(row=r_idx, column=2, value=f'=SUMIFS(\'Εκκαθαρίσεις ανά Ημερομηνία\'!M5:M{tot_r1-1}, \'Εκκαθαρίσεις ανά Ημερομηνία\'!C5:C{tot_r1-1}, "{m_code}", \'Εκκαθαρίσεις ανά Ημερομηνία\'!D5:D{tot_r1-1}, "ΣΥΝΕΡΓΑΤΗΣ") + IF(COUNTIFS(\'Εκκαθαρίσεις ανά Ημερομηνία\'!C5:C{tot_r1-1}, "{m_code}", \'Εκκαθαρίσεις ανά Ημερομηνία\'!D5:D{tot_r1-1}, "ΣΥΝΕΡΓΑΤΗΣ")=0, SUMIFS(\'Εκκαθαρίσεις ανά Ημερομηνία\'!M5:M{tot_r1-1}, \'Εκκαθαρίσεις ανά Ημερομηνία\'!C5:C{tot_r1-1}, "{m_code}"), 0)')
        p_cell.font = font_regular
        p_cell.number_format = "#,##0.00 €"
        p_cell.alignment = align_right
        p_cell.border = border_cell
        
        a_cell = ws2.cell(row=r_idx, column=3, value=f'=SUMIFS(\'Εκκαθαρίσεις ανά Ημερομηνία\'!P5:P{tot_r1-1}, \'Εκκαθαρίσεις ανά Ημερομηνία\'!C5:C{tot_r1-1}, "{m_code}", \'Εκκαθαρίσεις ανά Ημερομηνία\'!D5:D{tot_r1-1}, "ΣΥΝΕΡΓΑΤΗΣ")')
        a_cell.font = font_regular
        a_cell.number_format = "#,##0.00 €"
        a_cell.alignment = align_right
        a_cell.border = border_cell

        apct_cell = ws2.cell(row=r_idx, column=4, value=f'=IF(B{r_idx}>0, C{r_idx}/B{r_idx}, 0)')
        apct_cell.font = font_bold
        apct_cell.number_format = "0.00%"
        apct_cell.alignment = align_right
        apct_cell.border = border_cell

        g_cell = ws2.cell(row=r_idx, column=5, value=f'=SUMIFS(\'Εκκαθαρίσεις ανά Ημερομηνία\'!P5:P{tot_r1-1}, \'Εκκαθαρίσεις ανά Ημερομηνία\'!C5:C{tot_r1-1}, "{m_code}", \'Εκκαθαρίσεις ανά Ημερομηνία\'!D5:D{tot_r1-1}, "AGENCY")')
        g_cell.font = font_regular
        g_cell.number_format = "#,##0.00 €"
        g_cell.alignment = align_right
        g_cell.border = border_cell

        gpct_cell = ws2.cell(row=r_idx, column=6, value=f'=IF(B{r_idx}>0, E{r_idx}/B{r_idx}, 0)')
        gpct_cell.font = font_bold
        gpct_cell.number_format = "0.00%"
        gpct_cell.alignment = align_right
        gpct_cell.border = border_cell

        t_cell = ws2.cell(row=r_idx, column=7, value=f'=C{r_idx}+E{r_idx}')
        t_cell.font = font_bold
        t_cell.number_format = "#,##0.00 €"
        t_cell.alignment = align_right
        t_cell.border = border_cell

        tpct_cell = ws2.cell(row=r_idx, column=8, value=f'=IF(B{r_idx}>0, G{r_idx}/B{r_idx}, 0)')
        tpct_cell.font = font_bold
        tpct_cell.number_format = "0.00%"
        tpct_cell.alignment = align_right
        tpct_cell.border = border_cell

    tot_r2 = len(months_list) + 5
    ws2.row_dimensions[tot_r2].height = 24
    ws2.cell(row=tot_r2, column=1, value="ΣΥΝΟΛΑ:").font = font_bold
    ws2.cell(row=tot_r2, column=1).alignment = align_right
    ws2.cell(row=tot_r2, column=1).border = border_total

    for c_i, col_let in [(2, "B"), (3, "C"), (5, "E"), (7, "G")]:
        cell = ws2.cell(row=tot_r2, column=c_i, value=f"=SUM({col_let}5:{col_let}{tot_r2-1})")
        cell.font = font_bold
        cell.fill = fill_accent_gold
        cell.border = border_total
        cell.number_format = "#,##0.00 €"
        cell.alignment = align_right

    for c_i, num_let in [(4, "C"), (6, "E"), (8, "G")]:
        cell = ws2.cell(row=tot_r2, column=c_i, value=f"={num_let}{tot_r2}/B{tot_r2}")
        cell.font = font_bold
        cell.fill = fill_accent_gold
        cell.border = border_total
        cell.number_format = "0.00%"
        cell.alignment = align_right

    # SHEET 3: 📌 Ανά Συμβόλαιο & Μήνα (Per Policy & Month Detailed Sheet)
    ws3 = wb.create_sheet(title="Ανά Συμβόλαιο & Μήνα")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:K1")
    ws3["A1"] = "ERGO - ΑΝΑΛΥΣΗ ΑΜΟΙΒΩΝ & ΠΟΣΟΣΤΩΝ ΑΝΑ ΣΥΜΒΟΛΑΙΟ ΚΑΙ ΑΝΑ ΜΗΝΑ"
    ws3["A1"].font = font_title
    ws3["A1"].fill = fill_navy
    ws3["A1"].alignment = align_center
    ws3.row_dimensions[1].height = 36

    ws3.merge_cells("A2:K2")
    ws3["A2"] = "Πλήρης πίνακας δικαιωμάτων Συνεργάτη & Agency ανά συμβόλαιο για κάθε μήνα"
    ws3["A2"].font = font_subtitle
    ws3["A2"].fill = fill_navy
    ws3["A2"].alignment = align_center
    ws3.row_dimensions[2].height = 20

    headers3 = [
        "Μήνας", "Ημερομηνία", "Αρ. Συμβολαίου", "Ονοματεπώνυμο Πελάτη", "Έτος",
        "Καθαρά Ασφάλιστρα (€)", "Προμήθεια Συνεργάτη (€)", "Ποσοστό Συνεργάτη (%)",
        "Override Agency (€)", "Ποσοστό Agency (%)", "Σύνολο Αμοιβής (€)"
    ]
    ws3.row_dimensions[4].height = 28
    for c_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = border_header

    # Build rows per policy and per month
    policy_monthly_rows = []
    for m in months_list:
        m_df = merged_df[merged_df["Μήνας Εκκαθάρισης"] == m]
        unique_pols = m_df["Συμβόλαιο"].unique()
        for pol in unique_pols:
            p_df = m_df[m_df["Συμβόλαιο"] == pol]
            syn_r = p_df[p_df["Βαθμίδα"] == "ΣΥΝΕΡΓΑΤΗΣ"]
            agn_r = p_df[p_df["Βαθμίδα"] == "AGENCY"]
            
            date_val = p_df["Date_Obj"].iloc[0].strftime("%d/%m/%Y") if pd.notna(p_df["Date_Obj"].iloc[0]) else str(p_df["Εναρξη"].iloc[0])
            client_last = p_df["Επώνυμο"].iloc[0] if pd.notna(p_df["Επώνυμο"].iloc[0]) else ""
            client_first = p_df["Ονομα.1"].iloc[0] if pd.notna(p_df["Ονομα.1"].iloc[0]) else ""
            client_name = f"{client_last} {client_first}".strip()
            pol_year = p_df["Διαν.Ετος"].iloc[0]
            
            if len(syn_r) > 0:
                net_prem = syn_r["Καθαρά Σύνολο"].sum()
            else:
                net_prem = agn_r["Καθαρά Σύνολο"].sum()
                
            comm_syn = syn_r["Προμήθεια Σύνολο"].sum() if len(syn_r) > 0 else 0.0
            comm_agn = agn_r["Προμήθεια Σύνολο"].sum() if len(agn_r) > 0 else 0.0
            
            policy_monthly_rows.append((m, date_val, pol, client_name, pol_year, net_prem, comm_syn, comm_agn))

    for r_idx, row_tuple in enumerate(policy_monthly_rows, 5):
        ws3.row_dimensions[r_idx].height = 20
        m, date_val, pol, client_name, pol_year, net_prem, comm_syn, comm_agn = row_tuple
        
        ws3.cell(row=r_idx, column=1, value=m).alignment = align_center
        ws3.cell(row=r_idx, column=2, value=date_val).alignment = align_center
        ws3.cell(row=r_idx, column=3, value=str(pol)).alignment = align_center
        ws3.cell(row=r_idx, column=4, value=client_name).alignment = align_left
        ws3.cell(row=r_idx, column=5, value=pol_year).alignment = align_center
        
        p_cell = ws3.cell(row=r_idx, column=6, value=net_prem)
        p_cell.number_format = "#,##0.00 €"
        p_cell.alignment = align_right
        
        cs_cell = ws3.cell(row=r_idx, column=7, value=comm_syn)
        cs_cell.number_format = "#,##0.00 €"
        cs_cell.alignment = align_right

        pct_s = ws3.cell(row=r_idx, column=8, value=f'=IF(F{r_idx}>0, G{r_idx}/F{r_idx}, 0)')
        pct_s.number_format = "0.00%"
        pct_s.alignment = align_right

        ca_cell = ws3.cell(row=r_idx, column=9, value=comm_agn)
        ca_cell.number_format = "#,##0.00 €"
        ca_cell.alignment = align_right

        pct_a = ws3.cell(row=r_idx, column=10, value=f'=IF(F{r_idx}>0, I{r_idx}/F{r_idx}, 0)')
        pct_a.number_format = "0.00%"
        pct_a.alignment = align_right

        tot_c = ws3.cell(row=r_idx, column=11, value=f'=G{r_idx}+I{r_idx}')
        tot_c.font = font_bold
        tot_c.number_format = "#,##0.00 €"
        tot_c.alignment = align_right

        for c_i in range(1, 12):
            ws3.cell(row=r_idx, column=c_i).border = border_cell
            ws3.cell(row=r_idx, column=c_i).font = font_regular

    tot_r3 = len(policy_monthly_rows) + 5
    ws3.row_dimensions[tot_r3].height = 24
    ws3.cell(row=tot_r3, column=4, value="ΣΥΝΟΛΑ:").font = font_bold
    ws3.cell(row=tot_r3, column=4).alignment = align_right
    ws3.cell(row=tot_r3, column=4).border = border_total

    for c_i, col_let in [(6, "F"), (7, "G"), (9, "I"), (11, "K")]:
        cell = ws3.cell(row=tot_r3, column=c_i, value=f"=SUM({col_let}5:{col_let}{tot_r3-1})")
        cell.font = font_bold
        cell.fill = fill_accent_gold
        cell.border = border_total
        cell.number_format = "#,##0.00 €"
        cell.alignment = align_right

    pct_s_tot = ws3.cell(row=tot_r3, column=8, value=f"=G{tot_r3}/F{tot_r3}")
    pct_s_tot.font = font_bold
    pct_s_tot.fill = fill_accent_gold
    pct_s_tot.border = border_total
    pct_s_tot.number_format = "0.00%"
    pct_s_tot.alignment = align_right

    pct_a_tot = ws3.cell(row=tot_r3, column=10, value=f"=I{tot_r3}/F{tot_r3}")
    pct_a_tot.font = font_bold
    pct_a_tot.fill = fill_accent_gold
    pct_a_tot.border = border_total
    pct_a_tot.number_format = "0.00%"
    pct_a_tot.alignment = align_right

    # Column Widths
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 16
    ws1.column_dimensions['G'].width = 28
    ws1.column_dimensions['H'].width = 12
    ws1.column_dimensions['I'].width = 12
    ws1.column_dimensions['J'].width = 12
    ws1.column_dimensions['K'].width = 18
    ws1.column_dimensions['L'].width = 18
    ws1.column_dimensions['M'].width = 20
    ws1.column_dimensions['N'].width = 18
    ws1.column_dimensions['O'].width = 18
    ws1.column_dimensions['P'].width = 22
    ws1.column_dimensions['Q'].width = 14
    ws1.column_dimensions['R'].width = 22

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 24
    ws2.column_dimensions['C'].width = 26
    ws2.column_dimensions['D'].width = 22
    ws2.column_dimensions['E'].width = 26
    ws2.column_dimensions['F'].width = 22
    ws2.column_dimensions['G'].width = 24
    ws2.column_dimensions['H'].width = 22

    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 28
    ws3.column_dimensions['E'].width = 10
    ws3.column_dimensions['F'].width = 22
    ws3.column_dimensions['G'].width = 24
    ws3.column_dimensions['H'].width = 22
    ws3.column_dimensions['I'].width = 24
    ws3.column_dimensions['J'].width = 22
    ws3.column_dimensions['K'].width = 22

    out_file = r"g:\ΓΕΦΥΡΕΣ\ERGO ZWHS\ERGO_Health_Commissions_Consolidated_By_Date.xlsx"
    try:
        wb.save(out_file)
        print(f"Successfully updated consolidated file: {out_file}")
    except PermissionError:
        out_file_alt = r"g:\ΓΕΦΥΡΕΣ\ERGO ZWHS\ERGO_Health_Commissions_Consolidated_By_Date_v3.xlsx"
        wb.save(out_file_alt)
        print(f"File locked, saved to alternative file: {out_file_alt}")

if __name__ == "__main__":
    process_and_build()
