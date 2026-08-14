import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

def build_ergo_excel():
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Styles & Palette Definitions
    # Colors: Navy (#1B365D), Soft Blue (#E8EEF5), Accent Gold (#FFF3CD), Header Teal (#005F73), Light Gray (#F4F6F9), Border Gray (#D9D9D9)
    font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Calibri", size=11, italic=True, color="E8EEF5")
    font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    font_sub_header = Font(name="Calibri", size=11, bold=True, color="1B365D")
    font_bold = Font(name="Calibri", size=11, bold=True, color="000000")
    font_regular = Font(name="Calibri", size=11, color="000000")
    font_italic = Font(name="Calibri", size=10, italic=True, color="555555")

    fill_navy = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    fill_teal = PatternFill(start_color="005F73", end_color="005F73", fill_type="solid")
    fill_soft_blue = PatternFill(start_color="E8EEF5", end_color="E8EEF5", fill_type="solid")
    fill_accent_gold = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    fill_light_gray = PatternFill(start_color="F4F6F9", end_color="F4F6F9", fill_type="solid")
    fill_green_light = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thick_bottom_side = Side(border_style="medium", color="1B365D")
    double_bottom_side = Side(border_style="double", color="1B365D")

    border_cell = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    border_total = Border(top=thin_border_side, bottom=double_bottom_side)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center")

    # =========================================================================
    # SHEET 5: 5. Πίνακας Προμηθειών ΔΑΣ (Lookups & Master Data Table)
    # We build this first so lookup formulas in Sheet 1 can reference it cleanly.
    # =========================================================================
    ws5 = wb.create_sheet(title="5. Πίνακας Προμηθειών ΔΑΣ")
    ws5.views.sheetView[0].showGridLines = True

    # Title
    ws5.merge_cells("A1:J1")
    ws5["A1"] = "ERGO - ΠΙΝΑΚΑΣ ΠΡΟΜΗΘΕΙΩΝ ΚΑΝΟΝΙΣΜΟΥ ΠΩΛΗΣΕΩΝ ΔΑΣ (ΔΙΚΤΥΟ ΑΝΕΞΑΡΤΗΤΩΝ ΣΥΝΕΡΓΑΤΩΝ)"
    ws5["A1"].font = font_title
    ws5["A1"].fill = fill_navy
    ws5["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 40

    ws5.merge_cells("A2:J2")
    ws5["A2"] = "Ποσοστά προμηθειών πράκτορα ανά κατηγορία (Α, Β, Γ), έτος συμβολαίου & υπερπρομήθειες (overrides) γραφείου"
    ws5["A2"].font = font_subtitle
    ws5["A2"].fill = fill_navy
    ws5["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[2].height = 20

    headers5 = [
        "Κωδικός", "Κλάδος", "Ονομασία Προϊόντος / Παροχής", "Έτος",
        "Κατηγορία Α (%)", "Κατηγορία Β (%)", "Κατηγορία Γ (%)",
        "Override Unit Manager (%)", "Override Agency Manager (%)", "Σημειώσεις / Προϋποθέσεις"
    ]
    
    ws5.row_dimensions[4].height = 28
    for col_idx, h in enumerate(headers5, 1):
        cell = ws5.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = border_header

    # Commission Master Data Rows
    comm_data = [
        # Health Care
        ["020118", "Υγεία", "ERGO Health Care (Simple, Advanced, Superior)", "1ο Έτος", 0.2900, 0.3190, 0.3480, 0.10, 0.10, "Πρωτοετής προμήθεια υγείας"],
        ["020118", "Υγεία", "ERGO Health Care (Simple, Advanced, Superior)", "2ο Έτος", 0.2500, 0.2750, 0.3000, 0.10, 0.10, "Ανανέωση υγείας"],
        ["020118", "Υγεία", "ERGO Health Care (Simple, Advanced, Superior)", "3ο Έτος", 0.2500, 0.2750, 0.3000, 0.10, 0.10, "Ανανέωση υγείας"],
        ["020118", "Υγεία", "ERGO Health Care (Simple, Advanced, Superior)", "4ο-6ο Έτος", 0.2500, 0.2750, 0.3000, 0.10, 0.10, "Ανανέωση υγείας"],
        ["020118", "Υγεία", "ERGO Health Care (Simple, Advanced, Superior)", "7ο+ Έτος", 0.2500, 0.2750, 0.3000, 0.10, 0.10, "Ανανέωση υγείας"],
        
        # Best Health
        ["020119", "Υγεία", "ERGO Best Health", "1ο Έτος", 0.2600, 0.2860, 0.3120, 0.10, 0.10, "Πρωτοετής προμήθεια Best Health"],
        ["020119", "Υγεία", "ERGO Best Health", "2ο Έτος", 0.2300, 0.2530, 0.2760, 0.10, 0.10, "Ανανέωση Best Health"],
        ["020119", "Υγεία", "ERGO Best Health", "3ο Έτος", 0.2300, 0.2530, 0.2760, 0.10, 0.10, "Ανανέωση Best Health"],
        ["020119", "Υγεία", "ERGO Best Health", "4ο-6ο Έτος", 0.2300, 0.2530, 0.2760, 0.10, 0.10, "Ανανέωση Best Health"],
        ["020119", "Υγεία", "ERGO Best Health", "7ο+ Έτος", 0.2300, 0.2530, 0.2760, 0.10, 0.10, "Ανανέωση Best Health"],

        # Life - Ισόβια
        ["110118", "Ζωή", "ERGO Life - Ισόβια Ασφάλιση Θανάτου", "1ο Έτος", 0.5000, 0.5500, 0.6000, 0.10, 0.20, "1ο έτος Ισόβιας"],
        ["110118", "Ζωή", "ERGO Life - Ισόβια Ασφάλιση Θανάτου", "2ο Έτος", 0.1500, 0.1650, 0.1800, 0.10, 0.10, "2ο έτος Ισόβιας"],
        ["110118", "Ζωή", "ERGO Life - Ισόβια Ασφάλιση Θανάτου", "3ο Έτος", 0.0800, 0.0880, 0.0960, 0.10, 0.10, "3ο έτος Ισόβιας"],
        ["110118", "Ζωή", "ERGO Life - Ισόβια Ασφάλιση Θανάτου", "4ο-6ο Έτος", 0.0500, 0.0550, 0.0600, 0.10, 0.10, "4ο-6ο έτος Ισόβιας"],
        ["110118", "Ζωή", "ERGO Life - Ισόβια Ασφάλιση Θανάτου", "7ο+ Έτος", 0.0000, 0.0000, 0.0000, 0.00, 0.00, "7ο+ έτος Ισόβιας"],

        # Life - Πρόσκαιρη
        ["110318", "Ζωή", "ERGO Life - Πρόσκαιρη Ασφάλιση Θανάτου", "1ο Έτος", 0.2500, 0.2750, 0.3000, 0.10, 0.20, "1ο έτος Πρόσκαιρης"],
        ["110318", "Ζωή", "ERGO Life - Πρόσκαιρη Ασφάλιση Θανάτου", "2ο Έτος", 0.2000, 0.2200, 0.2400, 0.10, 0.10, "2ο έτος Πρόσκαιρης"],
        ["110318", "Ζωή", "ERGO Life - Πρόσκαιρη Ασφάλιση Θανάτου", "3ο Έτος", 0.1200, 0.1320, 0.1440, 0.10, 0.10, "3ο έτος Πρόσκαιρης"],
        ["110318", "Ζωή", "ERGO Life - Πρόσκαιρη Ασφάλιση Θανάτου", "4ο-6ο Έτος", 0.0600, 0.0660, 0.0720, 0.10, 0.10, "4ο-6ο έτος Πρόσκαιρης"],
        ["110318", "Ζωή", "ERGO Life - Πρόσκαιρη Ασφάλιση Θανάτου", "7ο+ Έτος", 0.0500, 0.0550, 0.0600, 0.10, 0.10, "7ο+ έτος Πρόσκαιρης"],

        # Savings - My Saving
        ["990119", "Αποταμίευση", "ERGO My Saving Simple & Junior", "1ο Έτος", 0.1458, 0.1604, 0.1750, 0.10, 0.10, "1ο έτος Αποταμίευσης"],
        ["990119", "Αποταμίευση", "ERGO My Saving Simple & Junior", "2ο Έτος", 0.0667, 0.0733, 0.0800, 0.10, 0.10, "2ο έτος Αποταμίευσης"],
        ["990119", "Αποταμίευση", "ERGO My Saving Simple & Junior", "3ο Έτος", 0.0417, 0.0458, 0.0500, 0.10, 0.10, "3ο έτος Αποταμίευσης"],
        ["990119", "Αποταμίευση", "ERGO My Saving Simple & Junior", "4ο-6ο Έτος", 0.0000, 0.0000, 0.0000, 0.00, 0.00, "4ο+ έτος"],
        ["990119", "Αποταμίευση", "ERGO My Saving Simple & Junior", "7ο+ Έτος", 0.0000, 0.0000, 0.0000, 0.00, 0.00, "7ο+ έτος"],

        # Unit-Linked - My Fund Flex Plan
        ["030122", "Unit-Linked", "ERGO My Fund Flex Plan", "1ο Έτος", 0.2208, 0.2429, 0.2650, 0.10, 0.10, "1ο έτος Flex Plan (€480-€750)"],
        ["030122", "Unit-Linked", "ERGO My Fund Flex Plan", "2ο Έτος", 0.1500, 0.1650, 0.1800, 0.10, 0.10, "2ο έτος Flex Plan"],
        ["030122", "Unit-Linked", "ERGO My Fund Flex Plan", "3ο Έτος", 0.0333, 0.0367, 0.0400, 0.10, 0.10, "3ο έτος Flex Plan"],
        ["030122", "Unit-Linked", "ERGO My Fund Flex Plan", "4ο-6ο Έτος", 0.0000, 0.0000, 0.0000, 0.00, 0.00, "4ο+ έτος"],
        ["030122", "Unit-Linked", "ERGO My Fund Flex Plan", "7ο+ Έτος", 0.0000, 0.0000, 0.0000, 0.00, 0.00, "7ο+ έτος"],

        # Unit-Linked - My Fund Invest Plan
        ["030222", "Unit-Linked", "ERGO My Fund Invest Plan", "1ο Έτος", 0.0142, 0.0156, 0.0170, 0.10, 0.10, "Εφάπαξ €4k-€10k"],
        ["030222", "Unit-Linked", "ERGO My Fund Invest Plan", "2ο Έτος", 0.0000, 0.0000, 0.0000, 0.00, 0.00, "Εφάπαξ - 2ο έτος"],
        ["030222", "Unit-Linked", "ERGO My Fund Invest Plan", "3ο Έτος", 0.0000, 0.0000, 0.0000, 0.00, 0.00, "Εφάπαξ - 3ο έτος"],
        ["030222", "Unit-Linked", "ERGO My Fund Invest Plan", "4ο-6ο Έτος", 0.0000, 0.0000, 0.0000, 0.00, 0.00, "Εφάπαξ - 4ο έτος"],
        ["030222", "Unit-Linked", "ERGO My Fund Invest Plan", "7ο+ Έτος", 0.0000, 0.0000, 0.0000, 0.00, 0.00, "Εφάπαξ - 7ο έτος"],

        # Group - My People Basic
        ["GROUP01", "Ομαδικά", "ERGO My People - Βασική Ασφάλιση Θανάτου", "1ο Έτος", 0.0800, 0.0800, 0.0800, 0.10, 0.10, "Βασική Ομαδική Θανάτου"],
        ["GROUP01", "Ομαδικά", "ERGO My People - Βασική Ασφάλιση Θανάτου", "2ο Έτος", 0.0800, 0.0800, 0.0800, 0.10, 0.10, "Ανανέωση Ομαδικού"],
        ["GROUP01", "Ομαδικά", "ERGO My People - Βασική Ασφάλιση Θανάτου", "3ο Έτος", 0.0800, 0.0800, 0.0800, 0.10, 0.10, "Ανανέωση Ομαδικού"],
        ["GROUP01", "Ομαδικά", "ERGO My People - Βασική Ασφάλιση Θανάτου", "4ο-6ο Έτος", 0.0800, 0.0800, 0.0800, 0.10, 0.10, "Ανανέωση Ομαδικού"],
        ["GROUP01", "Ομαδικά", "ERGO My People - Βασική Ασφάλιση Θανάτου", "7ο+ Έτος", 0.0800, 0.0800, 0.0800, 0.10, 0.10, "Ανανέωση Ομαδικού"],

        # Group - My People Riders
        ["GROUP02", "Ομαδικά", "ERGO My People - Συμπληρωματικές Καλύψεις", "1ο Έτος", 0.1500, 0.1500, 0.1500, 0.10, 0.10, "Συμπληρωματικά Ομαδικά"],
        ["GROUP02", "Ομαδικά", "ERGO My People - Συμπληρωματικές Καλύψεις", "2ο Έτος", 0.1500, 0.1500, 0.1500, 0.10, 0.10, "Ανανέωση Συμπληρωματικών"],
        ["GROUP02", "Ομαδικά", "ERGO My People - Συμπληρωματικές Καλύψεις", "3ο Έτος", 0.1500, 0.1500, 0.1500, 0.10, 0.10, "Ανανέωση Συμπληρωματικών"],
        ["GROUP02", "Ομαδικά", "ERGO My People - Συμπληρωματικές Καλύψεις", "4ο-6ο Έτος", 0.1500, 0.1500, 0.1500, 0.10, 0.10, "Ανανέωση Συμπληρωματικών"],
        ["GROUP02", "Ομαδικά", "ERGO My People - Συμπληρωματικές Καλύψεις", "7ο+ Έτος", 0.1500, 0.1500, 0.1500, 0.10, 0.10, "Ανανέωση Συμπληρωματικών"],
    ]

    for r_idx, row_val in enumerate(comm_data, 5):
        ws5.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(row_val, 1):
            cell = ws5.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_idx in [5, 6, 7, 8, 9]:
                cell.number_format = "0.00%"
                cell.alignment = align_right
            elif c_idx in [1, 4]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left


    # =========================================================================
    # SHEET 1: 🧮 1. Υπολογιστής & Καταχώρηση (Main Calculator Dashboard)
    # =========================================================================
    ws1 = wb.create_sheet(title="1. Υπολογιστής & Καταχώρηση")
    ws1.views.sheetView[0].showGridLines = True

    # Title Banner
    ws1.merge_cells("A1:G1")
    ws1["A1"] = "ERGO - ΥΠΟΛΟΓΙΣΤΗΣ ΠΡΟΜΗΘΕΙΩΝ, ΥΠΕΡΠΡΟΜΗΘΕΙΩΝ & ΚΑΛΥΨΕΩΝ ΣΥΜΒΟΛΑΙΟΥ"
    ws1["A1"].font = font_title
    ws1["A1"].fill = fill_navy
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:G2")
    ws1["A2"] = "Επιλέξτε παραμέτρους συμβολαίου για αυτόματο υπολογισμό δικαιωμάτων & προβολή ορίων καλύψεων"
    ws1["A2"].font = font_subtitle
    ws1["A2"].fill = fill_navy
    ws1["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 20

    # Section A: Input Parameters
    ws1.merge_cells("A4:C4")
    ws1["A4"] = "1. ΣТОΙΧΕΙΑ ΣΥΜΒΟΛΑΙΟΥ (ΕΙΣΑΓΩΓΗ ΧΡΗΣΤΗ)"
    ws1["A4"].font = font_sub_header
    ws1["A4"].fill = fill_soft_blue
    ws1["A4"].alignment = align_left
    ws1.row_dimensions[4].height = 24

    inputs = [
        ("B5", "A5", "Κλάδος Ασφάλισης:", "Υγεία"),
        ("B6", "A6", "Προϊόν / Πρόγραμμα:", "ERGO Health Care (Simple, Advanced, Superior)"),
        ("B7", "A7", "Έτος Συμβολαίου:", "1ο Έτος"),
        ("B8", "A8", "Κατηγορία Συνεργάτη:", "Κατηγορία Β"),
        ("B9", "A9", "Δομή / Ρόλος:", "Unit Manager (Δ')"),
        ("B10", "A10", "Καθαρά Ετήσια Ασφάλιστρα (€):", 1000.00)
    ]

    for cell_val, label_cell, label_text, default_val in inputs:
        row_num = int(label_cell[1:])
        ws1.row_dimensions[row_num].height = 22
        ws1[label_cell] = label_text
        ws1[label_cell].font = font_bold
        ws1[label_cell].alignment = align_left
        
        ws1.merge_cells(f"{cell_val}:{chr(ord(cell_val[0])+1)}{row_num}")
        input_target = ws1[cell_val]
        input_target.value = default_val
        input_target.font = font_bold
        input_target.fill = fill_accent_gold
        input_target.alignment = align_center
        input_target.border = border_cell
        if label_text.startswith("Καθαρά"):
            input_target.number_format = "#,##0.00 €"

    # Section B: Results (Formulas)
    ws1.merge_cells("E4:G4")
    ws1["E4"] = "2. ΑΠΟΤΕΛΕΣΜΑΤΑ ΥΠΟΛΟΓΙΣΜΟΥ (ΑΥТОΜΑΤΑ)"
    ws1["E4"].font = font_sub_header
    ws1["E4"].fill = fill_green_light
    ws1["E4"].alignment = align_left

    results = [
        ("G5", "E5", "Ποσοστό Προμήθειας Συνεργάτη (%):", '=SUMIFS(\'5. Πίνακας Προμηθειών ΔΑΣ\'!F5:F44, \'5. Πίνακας Προμηθειών ΔΑΣ\'!C5:C44, B6, \'5. Πίνακας Προμηθειών ΔΑΣ\'!D5:D44, B7)', "0.00%"),
        ("G6", "E6", "Ποσό Προμήθειας Συνεργάτη (€):", '=B10 * G5', "#,##0.00 €"),
        ("G7", "E7", "Υπερπρομήθεια / Override Γραφείου (%):", '=SUMIFS(\'5. Πίνακας Προμηθειών ΔΑΣ\'!H5:H44, \'5. Πίνακας Προμηθειών ΔΑΣ\'!C5:C44, B6, \'5. Πίνακας Προμηθειών ΔΑΣ\'!D5:D44, B7)', "0.00%"),
        ("G8", "E8", "Ποσό Υπερπρομήθειας Γραφείου (€):", '=B10 * G7', "#,##0.00 €"),
        ("G9", "E9", "Συνολικό Όφελος Γραφείου & Δικτύου (€):", '=G6 + G8', "#,##0.00 €"),
        ("G10", "E10", "Εκτιμώμενο Bonus Νέας Παραγωγής (%):", '=IF(B5="Υγεία", 0.03, IF(B5="Unit-Linked", 0.07, 0.00))', "0.00%")
    ]

    for cell_val, label_cell, label_text, formula_str, num_fmt in results:
        row_num = int(label_cell[1:])
        ws1.merge_cells(f"{label_cell}:{chr(ord(label_cell[0])+1)}{row_num}")
        ws1[label_cell] = label_text
        ws1[label_cell].font = font_bold
        ws1[label_cell].alignment = align_left

        res_target = ws1[cell_val]
        res_target.value = formula_str
        res_target.font = font_bold
        res_target.fill = fill_soft_blue
        res_target.alignment = align_right
        res_target.border = border_cell
        res_target.number_format = num_fmt

    # Section C: Dynamic Coverage Overview for Selected Product
    ws1.merge_cells("A13:G13")
    ws1["A13"] = "3. ΑΝΑΛΥΤΙΚΟΣ ΠΙΝΑΚΑΣ ΚΑΛΥΨΕΩΝ & ΠΑΡΟΧΩΝ ΕΠΙΛΕΓΜΕΝΟΥ ΣΥΜΒΟΛΑΙΟΥ"
    ws1["A13"].font = font_header
    ws1["A13"].fill = fill_teal
    ws1["A13"].alignment = align_center
    ws1.row_dimensions[13].height = 26

    cov_headers = ["Κατηγορία Κάλυψης", "Παροχή / Κάλυψη", "Simple / Βασική", "Advanced / Πλήρης", "Superior / Ανώτερη", "Απαλλαγή / Όρια", "Δίκτυο & Βοήθεια"]
    ws1.row_dimensions[14].height = 24
    for c_idx, h_text in enumerate(cov_headers, 1):
        cell = ws1.cell(row=14, column=c_idx, value=h_text)
        cell.font = font_header
        cell.fill = fill_navy
        cell.alignment = align_center
        cell.border = border_header

    cov_details = [
        ["Νοσηλεία", "Ανώτατο Όριο Κάλυψης ανά περιστατικό / έτος", "€300.000 / περιστατικό", "€300.000 / περιστατικό", "€500.000 / ασφαλιστικό έτος", "Ανάλογα με την επιλογή", "Παγκόσμια Κάλυψη"],
        ["Νοσηλεία", "Θέση Νοσηλείας σε Ιδιωτικά Νοσοκομεία", "Β' Θέση (Δίκλινο)", "Α' Θέση (Μονόκλινο)", "Α' Θέση (Μονόκλινο)", "Μείωση απαλλαγής σε χαμηλότερη θέση", "Ελλάδα & Εξωτερικό"],
        ["Νοσηλεία", "Αποκλειστικό Δίκτυο «4U» & Συμβεβλημένα", "100% Κάλυψη εντός 4U", "100% Κάλυψη εντός 4U", "100% Κάλυψη εντός 4U", "€0 / €500 / €1.500 / €3.000", "Άμεση κάλυψη"],
        ["Χειρουργική", "Αμοιβές Χειρουργού & Αναισθησιολόγου", "100% βάσει πίνακα", "100% βάσει πίνακα", "100% βάσει πίνακα", "Εντός ορίου νοσηλείας", "Ελεύθερη επιλογή"],
        ["ΜΕΘ", "Μονάδα Εντατικής Θεραπείας & Αυξημένης Φροντίδας", "100% Κάλυψη", "100% Κάλυψη", "100% Κάλυψη", "Χωρίς υπο-όριο", "Συμβεβλημένα & Μη"],
        ["Επείγοντα", "Επείγοντα Περιστατικά (Εξωτερικά Ιατρεία)", "€500 ανά περιστατικό", "€750 ανά περιστατικό", "€1.000 ανά περιστατικό", "Χωρίς απαλλαγή", "Δίκτυο 4U 24/7"],
        ["Διαγνωστικά", "Εξωνοσοκομειακές Διαγνωστικές Εξετάσεις", "Προαιρετική κάλυψη", "Προαιρετική κάλυψη", "Περιλαμβάνεται (€2.000)", "€0 συμμετοχή στο Δίκτυο", "Συμβεβλημένα Διαγνωστικά"],
        ["Επιδόματα", "Χειρουργικό & Νοσοκομειακό Επίδομα", "Παρέχεται", "Παρέχεται", "Παρέχεται", "Σε δημόσιο νοσοκομείο", "Έως €150/ημέρα"],
        ["Βοήθεια", "ERGOLIFE Assistance & Υγειονομική Μεταφορά", "Πλήρης Κάλυψη 24/7", "Πλήρης Κάλυψη 24/7", "Πλήρης Κάλυψη 24/7", "Χωρίς χρέωση", "Τηλ. 210 6505555"],
        ["Ζωή / Riders", "Θάνατος / Ατυχήματα / Ανικανότητα (Riders)", "Προαιρετικό κεφάλαιο", "Προαιρετικό κεφάλαιο", "Προαιρετικό κεφάλαιο", "Βάσει επιλογής λήπτη", "100% Αποζημίωση"],
        ["Αποταμίευση", "Εγγυημένο Κεφάλαιο & Απόδοση (My Saving)", "Εγγυημένο στη λήξη", "Εγγυημένο στη λήξη", "N/A", "Τακτικές καταβολές", "Ετήσια ενημέρωση"],
        ["Ομαδικά", "Εταιρική Κάλυψη Προσωπικού (My People)", "Βασική Θανάτου/ΜΑ", "Νοσοκομειακή/Εξωνοσοκ.", "Επιδόματα & Assistance", "Ελάχιστο 5 άτομα", "+5% Bonus Νέων"]
    ]

    for r_i, r_data in enumerate(cov_details, 15):
        ws1.row_dimensions[r_i].height = 20
        for c_i, val in enumerate(r_data, 1):
            cell = ws1.cell(row=r_i, column=c_i, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_i == 1:
                cell.font = font_bold
                cell.fill = fill_light_gray
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Add Data Validation Dropdowns for Sheet 1 inputs
    dv_branch = DataValidation(type="list", formula1='"Υγεία,Ζωή,Αποταμίευση,Unit-Linked,Ομαδικά"', allow_blank=False)
    dv_product = DataValidation(type="list", formula1='"ERGO Health Care (Simple, Advanced, Superior),ERGO Best Health,ERGO Life - Ισόβια Ασφάλιση Θανάτου,ERGO Life - Πρόσκαιρη Ασφάλιση Θανάτου,ERGO My Saving Simple & Junior,ERGO My Fund Flex Plan,ERGO My Fund Invest Plan,ERGO My People - Βασική Ασφάλιση Θανάτου,ERGO My People - Συμπληρωματικές Καλύψεις"', allow_blank=False)
    dv_year = DataValidation(type="list", formula1='"1ο Έτος,2ο Έτος,3ο Έτος,4ο-6ο Έτος,7ο+ Έτος"', allow_blank=False)
    dv_cat = DataValidation(type="list", formula1='"Κατηγορία Α,Κατηγορία Β,Κατηγορία Γ"', allow_blank=False)
    dv_role = DataValidation(type="list", formula1='"Απευθείας Πράκτορας,Unit Manager (Δ\'),Agency Manager (Β\')"', allow_blank=False)

    ws1.add_data_validation(dv_branch)
    ws1.add_data_validation(dv_product)
    ws1.add_data_validation(dv_year)
    ws1.add_data_validation(dv_cat)
    ws1.add_data_validation(dv_role)

    dv_branch.add("B5")
    dv_product.add("B6")
    dv_year.add("B7")
    dv_cat.add("B8")
    dv_role.add("B9")


    # =========================================================================
    # SHEET 2: 🏥 2. Καλύψεις Υγείας (Detailed Health Matrix)
    # =========================================================================
    ws2 = wb.create_sheet(title="2. Καλύψεις Υγείας")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:F1")
    ws2["A1"] = "ERGO HEALTH CARE & BEST HEALTH - ΑΝΑΛΥΤΙΚΟΣ ΠΙΝΑΚΑΣ ΚΑΛΥΨΕΩΝ ΥΓΕΙΑΣ"
    ws2["A1"].font = font_title
    ws2["A1"].fill = fill_navy
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 36

    ws2.merge_cells("A2:F2")
    ws2["A2"] = "Συγκριτικός πίνακας ορίων, θέσεων νοσηλείας, δικτύων και παροχών ατομικών προγραμμάτων υγείας"
    ws2["A2"].font = font_subtitle
    ws2["A2"].fill = fill_navy
    ws2["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[2].height = 20

    headers2 = ["Ασφαλιστική Κάλυψη / Παροχή", "Health Care Simple", "Health Care Advanced", "Health Care Superior", "ERGO Best Health", "Σημειώσεις & Ειδικοί Όροι"]
    ws2.row_dimensions[4].height = 26
    for c_idx, h in enumerate(headers2, 1):
        cell = ws2.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = border_header

    health_matrix = [
        ["Ανώτατο Όριο Κάλυψης", "€300.000 ανά περιστατικό", "€300.000 ανά περιστατικό", "€500.000 ανά ασφαλιστικό έτος", "€1.000.000 ανά έτος", "Ανανεώνεται ετησίως"],
        ["Θέση Νοσηλείας", "Β' Θέση (Δίκλινο)", "Α' Θέση (Μονόκλινο)", "Α' Θέση (Μονόκλινο)", "Α' Θέση (Μονόκλινο)", "Δυνατότητα νοσηλείας σε χαμηλότερη θέση"],
        ["Δίκτυο Νοσοκομείων «4U»", "Αποκλειστική πρόσβαση 4U", "Αναγνωρισμένα & 4U", "Αναγνωρισμένα & 4U", "Πλήρες Δίκτυο ERGO", "100% απευθείας εξόφληση"],
        ["Επιλογές Απαλλαγής", "€0 / €500 / €1.500 / €3.000", "€0 / €500 / €1.500 / €3.000", "€0 / €500 / €1.500 / €3.000", "€1.500 / €3.000", "Επιλέγεται κατά την έκδοση"],
        ["Ποσοστό Κάλυψης Εντός Δικτύου", "100% μετά την απαλλαγή", "100% μετά την απαλλαγή", "100% μετά την απαλλαγή", "100% μετά την απαλλαγή", "Απευθείας πληρωμή νοσηλίων"],
        ["Ποσοστό Κάλυψης Εκτός Δικτύου", "80% (Ελλάδα), 70% (Εξωτερικό)", "80% (Ελλάδα), 80% (Εξωτερικό)", "90% (Ελλάδα), 80% (Εξωτερικό)", "80% (Παγκόσμια)", "Απολογιστική εξόφληση"],
        ["Χειρουργικά Έξοδα & Ιατροί", "100% βάσει πίνακα αμοιβών", "100% βάσει πίνακα αμοιβών", "100% βάσει πίνακα αμοιβών", "100% βάσει πίνακα", "Χειρουργός & Αναισθησιολόγος"],
        ["Μονάδα Εντατικής Θεραπείας (ΜΕΘ)", "100% Κάλυψη", "100% Κάλυψη", "100% Κάλυψη", "100% Κάλυψη", "Χωρίς υπο-όριο ημερών"],
        ["Επείγοντα Περιστατικά", "Έως €500 / περιστατικό", "Έως €750 / περιστατικό", "Έως €1.000 / περιστατικό", "Έως €1.000 / περιστατικό", "24/7 στα εξωτερικά ιατρεία"],
        ["Εξωνοσοκομειακές Εξετάσεις", "Προαιρετικό πακέτο", "Προαιρετικό πακέτο", "Περιλαμβάνεται (€2.000)", "Περιλαμβάνεται", "Σε συμβεβλημένα διαγνωστικά"],
        ["Χειρουργικό Επίδομα", "Έως €2.500", "Έως €3.500", "Έως €5.000", "Έως €5.000", "Σε περίπτωση μη υποβολής εξόδων"],
        ["Νοσοκομειακό Επίδομα", "€100/ημέρα (έως 30 ημέρες)", "€120/ημέρα (έως 30 ημέρες)", "€150/ημέρα (έως 30 ημέρες)", "€150/ημέρα", "Για νοσηλεία σε δημόσιο"],
        ["Χημειοθεραπείες / Ακτινοθεραπείες", "100% Κάλυψη", "100% Κάλυψη", "100% Κάλυψη", "100% Κάλυψη", "Εντός ή εκτός νοσοκομείου"],
        ["ERGOLIFE Assistance", "Πλήρης 24/7 Υγειονομική", "Πλήρης 24/7 Υγειονομική", "Πλήρης 24/7 Υγειονομική", "Πλήρης 24/7 Υγειονομική", "Ασθενοφόρο & ιατρικές συμβουλές"]
    ]

    for r_idx, r_val in enumerate(health_matrix, 5):
        ws2.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(r_val, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_idx == 1:
                cell.font = font_bold
                cell.fill = fill_light_gray
                cell.alignment = align_left
            else:
                cell.alignment = align_center if c_idx < 6 else align_left


    # =========================================================================
    # SHEET 3: 🛡️ 3. Καλύψεις Ζωής & Αποταμίευσης
    # =========================================================================
    ws3 = wb.create_sheet(title="3. Καλύψεις Ζωής & Αποταμίευσης")
    ws3.views.sheetView[0].showGridLines = True

    ws3.merge_cells("A1:E1")
    ws3["A1"] = "ERGO LIFE, MY SAVING & MY FUND - ΑΝΑΛΥΤΙΚΕΣ ΚΑΛΥΨΕΙΣ ΖΩΗΣ & ΕΠΕΝΔΥΣΕΩΝ"
    ws3["A1"].font = font_title
    ws3["A1"].fill = fill_navy
    ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 36

    ws3.merge_cells("A2:E2")
    ws3["A2"] = "Βασικές ασφαλίσεις θανάτου, ταμιευτήρια, Unit-Linked & συμπληρωματικές προστασίες (Riders)"
    ws3["A2"].font = font_subtitle
    ws3["A2"].fill = fill_navy
    ws3["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[2].height = 20

    headers3 = ["Κατηγορία Προϊόντος / Rider", "Ονομασία Παροχής", "Κωδικός", "Περιγραφή & Όρια Κάλυψης", "Προϋποθέσεις & Ηλικιακά Όρια"]
    ws3.row_dimensions[4].height = 26
    for c_idx, h in enumerate(headers3, 1):
        cell = ws3.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = border_header

    life_matrix = [
        ["Βασική Ζωής", "Ισόβια Ασφάλιση Θανάτου", "110118", "Καταβολή ασφαλισμένου κεφαλαίου στους δικαιούχους σε περίπτωση απώλειας ζωής", "Είσοδος 18-65 ετών"],
        ["Βασική Ζωής", "Πρόσκαιρη Ασφάλιση Θανάτου", "110318", "Κάλυψη θανάτου για ορισμένη χρονική διάρκεια (1-30 έτη)", "Είσοδος 18-65 ετών"],
        ["Αποταμίευση", "ERGO My Saving Simple", "990119", "Εγγυημένο κεφάλαιο στη λήξη + δυνατότητα υπεραπόδοσης. Περιοδικό ασφάλιστρο", "Ελάχιστη διάρκεια 10 έτη"],
        ["Αποταμίευση", "ERGO My Saving Junior", "990219", "Παιδικό αποταμιευτικό πρόγραμμα εξασφάλισης σπουδών/κεφαλαίου", "Για παιδιά 0-14 ετών"],
        ["Unit-Linked", "ERGO My Fund Flex Plan", "030122", "Επενδυτικό πρόγραμμα περιοδικών καταβολών συνδεδεμένο με αμοιβαία κεφάλαια", "Ελεύθερη επιλογή χαρτοφυλακίου"],
        ["Unit-Linked", "ERGO My Fund Invest Plan", "030222", "Επενδυτικό πρόγραμμα εφάπαξ καταβολής (Invest Plan)", "Ελάχιστο εφάπαξ €4.000"],
        ["Συμπληρωματική (Rider)", "Θάνατος από Ατύχημα", "130118", "Διπλασιασμός κεφαλαίου σε περίπτωση θανάτου από ατύχημα", "Έως 65 ετών"],
        ["Συμπληρωματική (Rider)", "Μόνιμη Ολική / Μερική Ανικανότητα", "130218", "Καταβολή ποσοστού κεφαλαίου αναλόγως του βαθμού αναπηρίας", "Βάσει πίνακα ανικανότητας"],
        ["Συμπληρωματική (Rider)", "Ιατροφαρμακευτικά Έξοδα Ατυχήματος", "130418", "Κάλυψη εξόδων θεραπείας από ατύχημα έως το όριο", "Έως €3.000 / ατύχημα"],
        ["Συμπληρωματική (Rider)", "Απώλεια Εισοδήματος", "130718", "Μηνιαίο επίδομα λόγω απουσίας από την εργασία από ασθένεια/ατύχημα", "Έως 24 μήνες"],
        ["Συμπληρωματική (Rider)", "Απαλλαγή Πληρωμής Ασφαλίστρων (ΑΠΑ)", "130518", "Η εταιρεία αναλαμβάνει την πληρωμή ασφαλίστρων σε διαρκή ανικανότητα", "Προστασία συμβολαίου"],
        ["Συμπληρωματική (Rider)", "Σοβαρές Ασθένειες", "021722", "Καταβολή εφάπαξ κεφαλαίου σε διάγνωση σοβαρής ασθένειας (καρκίνος, έμφραγμα κλπ)", "12 ή 31 ασθένειες"]
    ]

    for r_idx, r_val in enumerate(life_matrix, 5):
        ws3.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(r_val, 1):
            cell = ws3.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_idx in [1, 3]:
                cell.font = font_bold
                cell.fill = fill_light_gray
                cell.alignment = align_center
            else:
                cell.alignment = align_left


    # =========================================================================
    # SHEET 4: 👥 4. Καλύψεις Ομαδικών
    # =========================================================================
    ws4 = wb.create_sheet(title="4. Καλύψεις Ομαδικών")
    ws4.views.sheetView[0].showGridLines = True

    ws4.merge_cells("A1:E1")
    ws4["A1"] = "ERGO MY PEOPLE - ΟΜΑΔΙΚΕΣ ΑΣΦΑΛΙΣΕΙΣ ΠΡΟΣΩΠΙΚΟΥ ΕΠΙΧΕΙΡΗΣΕΩΝ"
    ws4["A1"].font = font_title
    ws4["A1"].fill = fill_navy
    ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 36

    ws4.merge_cells("A2:E2")
    ws4["A2"] = "Προγράμματα ομαδικής ασφάλισης Ζωής, Ανικανότητας, Νοσοκομειακής & Εξωνοσοκομειακής Περίθαλψης"
    ws4["A2"].font = font_subtitle
    ws4["A2"].fill = fill_navy
    ws4["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[2].height = 20

    headers4 = ["Κατηγορία Κάλυψης", "Παροχή Ομαδικού Προγράμματος", "Περιγραφή & Όρια Κάλυψης", "Προϋποθέσεις & Δικαιολογητικά", "Προμήθειες & Bonus"]
    ws4.row_dimensions[4].height = 26
    for c_idx, h in enumerate(headers4, 1):
        cell = ws4.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = border_header

    group_matrix = [
        ["Βασική Ασφάλιση", "Πρόσκαιρη Ασφάλιση Θανάτου", "Κεφάλαιο ασφάλισης ανά εργαζόμενο (σταθερό ή πολλαπλάσιο μισθού)", "Ελάχιστο 5 ασφαλισμένα άτομα", "8.00% προμήθεια (όλα τα έτη)"],
        ["Συμπληρωματική", "Μόνιμη Ολική / Μερική Ανικανότητα (ΜΟΑ/ΜΜΑ)", "Καταβολή κεφαλαίου σε περίπτωση ατυχήματος ή ασθένειας", "Συνδέεται με τη βασική κάλυψη", "15.00% προμήθεια (όλα τα έτη)"],
        ["Συμπληρωματική", "Ευρεία Νοσοκομειακή Περίθαλψη", "Κάλυψη εξόδων νοσηλείας σε ιδιωτικά & δημόσια νοσοκομεία", "Ετήσιο όριο ανά ασφαλισμένο (€10k-€50k)", "15.00% προμήθεια (όλα τα έτη)"],
        ["Συμπληρωματική", "Εξωνοσοκομειακή Περίθαλψη", "Διαγνωστικές εξετάσεις, ιατρικές επισκέψεις, φάρμακα", "Σε συμβεβλημένο δίκτυο", "15.00% προμήθεια (όλα τα έτη)"],
        ["Συμπληρωματική", "Νοσοκομειακό & Χειρουργικό Επίδομα", "Ημερήσιο επίδομα για κάθε ημέρα νοσηλείας", "Έως 180 ημέρες ανά έτος", "15.00% προμήθεια (όλα τα έτη)"],
        ["Υπηρεσία", "ERGOLIFE Assistance 24/7", "Υγειονομική μεταφορά, ασθενοφόρο, τηλεφωνική υποστήριξη", "Περιλαμβάνεται σε όλα τα συμβόλαια", "Χωρίς πρόσθετη χρέωση"],
        ["Bonus Νέων", "Bonus Πρωτοετούς Παραγωγής Νέων Ομαδικών", "Έκτακτο ετήσιο bonus +5.00% επί των εισπραγμένων ασφαλίστρων", "Πρωτοασφαλιζόμενη επιχείρηση (έως 100 άτομα)", "+5.00% επιπλέον αμοιβή"]
    ]

    for r_idx, r_val in enumerate(group_matrix, 5):
        ws4.row_dimensions[r_idx].height = 22
        for c_idx, val in enumerate(r_val, 1):
            cell = ws4.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_idx == 1:
                cell.font = font_bold
                cell.fill = fill_light_gray
                cell.alignment = align_center
            else:
                cell.alignment = align_left


    # =========================================================================
    # SHEET 6: 🏆 6. Πριμ & Bonus
    # =========================================================================
    ws6 = wb.create_sheet(title="6. Πριμ & Bonus")
    ws6.views.sheetView[0].showGridLines = True

    ws6.merge_cells("A1:E1")
    ws6["A1"] = "ERGO - ΚΛΙΜΑΚΕΣ ΠΡΙΜ & BONUS ΠΑΡΑΓΩΓΗΣ ΚΑΝΟΝΙΣΜΟΥ ΠΩΛΗΣΕΩΝ"
    ws6["A1"].font = font_title
    ws6["A1"].fill = fill_navy
    ws6["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[1].height = 36

    ws6.merge_cells("A2:E2")
    ws6["A2"] = "Ετήσια πρόσθετα κίνητρα διατήρησης, καλού αποτελέσματος (ΔΖ) & νέας παραγωγής"
    ws6["A2"].font = font_subtitle
    ws6["A2"].fill = fill_navy
    ws6["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws6.row_dimensions[2].height = 20

    # 1. Bonus Διατήρησης
    ws6.merge_cells("A4:E4")
    ws6["A4"] = "1. BONUS ΔΙΑΤΗΡΗΣΗΣ & ΑΝΑΠΤΥΞΗΣ ΧΑΡТОΦΥΛΑΚΙΟΥ (ΖΩΗΣ & ΥΓΕΙΑΣ)"
    ws6["A4"].font = font_sub_header
    ws6["A4"].fill = fill_soft_blue

    headers6_1 = ["Ύψος Χαρτοφυλακίου (€)", "Ποσοστό Bonus (%)", "Προϋποθέσεις", "Πεδίο Εφαρμογής", "Χρόνος Καταβολής"]
    ws6.row_dimensions[5].height = 24
    for c_idx, h in enumerate(headers6_1, 1):
        cell = ws6.cell(row=5, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center

    b1_data = [
        ["> €50.000", 0.0100, "Ελάχιστο χαρτοφυλάκιο €50.000", "Ετήσια ασφάλιστρα Ζωής & Υγείας", "Ετησίως (Απρίλιος)"],
        ["> €100.000", 0.0150, "Ελάχιστο χαρτοφυλάκιο €100.000", "Ετήσια ασφάλιστρα Ζωής & Υγείας", "Ετησίως (Απρίλιος)"],
        ["> €150.000", 0.0200, "Ελάχιστο χαρτοφυλάκιο €150.000", "Ετήσια ασφάλιστρα Ζωής & Υγείας", "Ετησίως (Απρίλιος)"]
    ]
    for r_idx, r_val in enumerate(b1_data, 6):
        ws6.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(r_val, 1):
            cell = ws6.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_idx == 2:
                cell.number_format = "0.00%"
                cell.alignment = align_right
            else:
                cell.alignment = align_center if c_idx == 1 else align_left

    # 2. Bonus Καλού Αποτελέσματος
    ws6.merge_cells("A10:E10")
    ws6["A10"] = "2. BONUS ΚΑΛΟΥ ΑΠΟΤΕΛΕΣΜΑΤΟΣ / ΔΕΙKTH ΖΗΜΙΩΝ (ΔΖ ΥΓΕΙΑΣ)"
    ws6["A10"].font = font_sub_header
    ws6["A10"].fill = fill_soft_blue

    headers6_2 = ["Δείκτης Ζημιών (ΔΖ)", "Ποσοστό Bonus (%)", "Ελάχιστη Παραγωγή", "Υπολογισμός", "Παρατηρήσεις"]
    ws6.row_dimensions[11].height = 24
    for c_idx, h in enumerate(headers6_2, 1):
        cell = ws6.cell(row=11, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center

    b2_data = [
        ["ΔΖ < 50%", 0.0200, "€30.000 εισπραγμένη παραγωγής", "Αποζημιώσεις / Δεδουλευμένα Ασφάλιστρα", "Εξαιρείται η Πρωτοβάθμια"],
        ["ΔΖ < 45%", 0.0300, "€30.000 εισπραγμένη παραγωγής", "Αποζημιώσεις / Δεδουλευμένα Ασφάλιστρα", "Εξαιρείται η Πρωτοβάθμια"],
        ["ΔΖ < 40%", 0.0400, "€30.000 εισπραγμένη παραγωγής", "Αποζημιώσεις / Δεδουλευμένα Ασφάλιστρα", "Εξαιρείται η Πρωτοβάθμια"],
        ["ΔΖ < 35%", 0.0500, "€30.000 εισπραγμένη παραγωγής", "Αποζημιώσεις / Δεδουλευμένα Ασφάλιστρα", "Εξαιρείται η Πρωτοβάθμια"],
        ["ΔΖ < 30%", 0.0600, "€30.000 εισπραγμένη παραγωγής", "Αποζημιώσεις / Δεδουλευμένα Ασφάλιστρα", "Εξαιρείται η Πρωτοβάθμια"]
    ]
    for r_idx, r_val in enumerate(b2_data, 12):
        ws6.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(r_val, 1):
            cell = ws6.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_idx == 2:
                cell.number_format = "0.00%"
                cell.alignment = align_right
            else:
                cell.alignment = align_center if c_idx == 1 else align_left

    # 3. Bonus Νέας Παραγωγής Υγείας
    ws6.merge_cells("A18:E18")
    ws6["A18"] = "3. BONUS ΠΡΩТОΕΤΟΥΣ ΝΕΑΣ ΠΑΡΑΓΩΓΗΣ ΑΣΦΑΛΙΣΕΩΝ ΥΓΕΙΑΣ"
    ws6["A18"].font = font_sub_header
    ws6["A18"].fill = fill_soft_blue

    headers6_3 = ["Εισπραγμένη Νέα Παραγωγή", "Ποσοστό Bonus (%)", "Εφαρμογή", "Εξαιρέσεις", "Κατηγορίες"]
    ws6.row_dimensions[19].height = 24
    for c_idx, h in enumerate(headers6_3, 1):
        cell = ws6.cell(row=19, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center

    b3_data = [
        ["> €10.000", 0.0100, "Στο σύνολο της νέας παραγωγής", "Δεν προσμετρώνται αντικαταστάσεις", "Ισχύει για Κατηγορίες Α, Β, Γ"],
        ["> €20.000", 0.0200, "Στο σύνολο της νέας παραγωγής", "Δεν προσμετρώνται αντικαταστάσεις", "Ισχύει για Κατηγορίες Α, Β, Γ"],
        ["> €30.000", 0.0300, "Στο σύνολο της νέας παραγωγής", "Δεν προσμετρώνται αντικαταστάσεις", "Ισχύει για Κατηγορίες Α, Β, Γ"],
        ["> €40.000", 0.0400, "Στο σύνολο της νέας παραγωγής", "Δεν προσμετρώνται αντικαταστάσεις", "Ισχύει για Κατηγορίες Α, Β, Γ"],
        ["> €50.000", 0.0500, "Στο σύνολο της νέας παραγωγής", "Δεν προσμετρώνται αντικαταστάσεις", "Ισχύει για Κατηγορίες Α, Β, Γ"]
    ]
    for r_idx, r_val in enumerate(b3_data, 20):
        ws6.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(r_val, 1):
            cell = ws6.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_idx == 2:
                cell.number_format = "0.00%"
                cell.alignment = align_right
            else:
                cell.alignment = align_center if c_idx == 1 else align_left


    # =========================================================================
    # SHEET 7: 📋 7. Αρχείο Συμβολαίων (Contract Ledger Log)
    # =========================================================================
    ws7 = wb.create_sheet(title="7. Αρχείο Συμβολαίων")
    ws7.views.sheetView[0].showGridLines = True

    ws7.merge_cells("A1:O1")
    ws7["A1"] = "ERGO - ΑΡΧΕΙΟ ΚΑΤΑΧΩΡΗΣΗΣ & ΠΑΡΑΚΟΛΟΥΘΗΣΗΣ ΣΥΜΒΟΛΑΙΩΝ"
    ws7["A1"].font = font_title
    ws7["A1"].fill = fill_navy
    ws7["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws7.row_dimensions[1].height = 36

    ws7.merge_cells("A2:O2")
    ws7["A2"] = "Φόρμα παρακολούθησης νέων συμβολαίων με αυτόματο υπολογισμό δικαιωμάτων προμηθειών & υπερπρομηθειών"
    ws7["A2"].font = font_subtitle
    ws7["A2"].fill = fill_navy
    ws7["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws7.row_dimensions[2].height = 20

    headers7 = [
        "Α/Α", "Ημερομηνία", "Αρ. Συμβολαίου", "Ονοματεπώνυμο Λήπτη", "Κλάδος",
        "Προϊόν / Πρόγραμμα", "Έτος", "Κατηγορία", "Καθαρά Ασφάλιστρα (€)",
        "Προμήθεια (%)", "Προμήθεια (€)", "Override (%)", "Override (€)",
        "Συνολικό Όφελος (€)", "Σημειώσεις"
    ]
    ws7.row_dimensions[4].height = 26
    for c_idx, h in enumerate(headers7, 1):
        cell = ws7.cell(row=4, column=c_idx, value=h)
        cell.font = font_header
        cell.fill = fill_teal
        cell.alignment = align_center
        cell.border = border_header

    # Pre-populate sample contract entries
    sample_ledger = [
        [1, "2026-08-01", "ERG-99401", "Γεώργιος Παπαδόπουλος", "Υγεία", "ERGO Health Care (Simple, Advanced, Superior)", "1ο Έτος", "Κατηγορία Β", 1200.00, 0.3190, "=I5*J5", 0.1000, "=I5*L5", "=K5+M5", "Superior Α' Θέση"],
        [2, "2026-08-03", "ERG-99402", "Μαρία Κωνσταντίνου", "Ζωή", "ERGO Life - Ισόβια Ασφάλιση Θανάτου", "1ο Έτος", "Κατηγορία Β", 850.00, 0.5500, "=I6*J6", 0.1000, "=I6*L6", "=K6+M6", "Ισόβια + Θάνατος Ατυχήματος"],
        [3, "2026-08-05", "ERG-99403", "Αλέξανδρος Νικολάου", "Αποταμίευση", "ERGO My Saving Simple & Junior", "1ο Έτος", "Κατηγορία Β", 1500.00, 0.1604, "=I7*J7", 0.1000, "=I7*L7", "=K7+M7", "My Saving Simple 15ετία"],
        [4, "2026-08-08", "ERG-99404", "Tech Solutions Ε.Π.Ε.", "Ομαδικά", "ERGO My People - Βασική Ασφάλιση Θανάτου", "1ο Έτος", "Κατηγορία Β", 4500.00, 0.0800, "=I8*J8", 0.1000, "=I8*L8", "=K8+M8", "Ομαδικό 25 άτομα + Bonus 5%"],
        [5, "2026-08-10", "ERG-99405", "Ελένη Δημητρίου", "Υγεία", "ERGO Health Care (Simple, Advanced, Superior)", "2ο Έτος", "Κατηγορία Γ", 950.00, 0.3000, "=I9*J9", 0.1000, "=I9*L9", "=K9+M9", "Ανανέωση 2ου έτους Advanced"]
    ]

    for r_idx, r_val in enumerate(sample_ledger, 5):
        ws7.row_dimensions[r_idx].height = 20
        for c_idx, val in enumerate(r_val, 1):
            cell = ws7.cell(row=r_idx, column=c_idx, value=val)
            cell.font = font_regular
            cell.border = border_cell
            if c_idx in [9, 11, 13, 14]:
                cell.number_format = "#,##0.00 €"
                cell.alignment = align_right
            elif c_idx in [10, 12]:
                cell.number_format = "0.00%"
                cell.alignment = align_right
            elif c_idx in [1, 2, 3, 5, 7, 8]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    # Total Summary Row for Ledger
    tot_row = len(sample_ledger) + 5
    ws7.row_dimensions[tot_row].height = 24
    ws7.cell(row=tot_row, column=4, value="ΣΥΝΟΛΑ:").font = font_bold
    ws7.cell(row=tot_row, column=4).alignment = align_right

    for c_i, col_let in [(9, "I"), (11, "K"), (13, "M"), (14, "N")]:
        cell = ws7.cell(row=tot_row, column=c_i, value=f"=SUM({col_let}5:{col_let}{tot_row-1})")
        cell.font = font_bold
        cell.fill = fill_accent_gold
        cell.border = border_total
        cell.number_format = "#,##0.00 €"
        cell.alignment = align_right


    # =========================================================================
    # Auto-adjust column widths across all worksheets
    # =========================================================================
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.number_format and '%' in cell.number_format:
                    val_str += ' %'
                if len(val_str) > max_len and not cell.coordinate in ['A1', 'A2', 'A4']:
                    max_len = len(val_str)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 14)

    # Specific fine-tuning widths for best visual representation
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 22
    ws1.column_dimensions['E'].width = 36
    ws1.column_dimensions['F'].width = 22
    ws1.column_dimensions['G'].width = 26

    ws2.column_dimensions['A'].width = 32
    ws2.column_dimensions['B'].width = 26
    ws2.column_dimensions['C'].width = 26
    ws2.column_dimensions['D'].width = 28
    ws2.column_dimensions['E'].width = 24
    ws2.column_dimensions['F'].width = 30

    ws3.column_dimensions['A'].width = 24
    ws3.column_dimensions['B'].width = 34
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 50
    ws3.column_dimensions['E'].width = 30

    ws4.column_dimensions['A'].width = 24
    ws4.column_dimensions['B'].width = 34
    ws4.column_dimensions['C'].width = 45
    ws4.column_dimensions['D'].width = 35
    ws4.column_dimensions['E'].width = 28

    ws5.column_dimensions['A'].width = 12
    ws5.column_dimensions['B'].width = 16
    ws5.column_dimensions['C'].width = 44
    ws5.column_dimensions['D'].width = 14
    ws5.column_dimensions['E'].width = 18
    ws5.column_dimensions['F'].width = 18
    ws5.column_dimensions['G'].width = 18
    ws5.column_dimensions['H'].width = 22
    ws5.column_dimensions['I'].width = 22
    ws5.column_dimensions['J'].width = 35

    ws7.column_dimensions['A'].width = 8
    ws7.column_dimensions['B'].width = 14
    ws7.column_dimensions['C'].width = 16
    ws7.column_dimensions['D'].width = 28
    ws7.column_dimensions['E'].width = 16
    ws7.column_dimensions['F'].width = 40
    ws7.column_dimensions['G'].width = 12
    ws7.column_dimensions['H'].width = 16
    ws7.column_dimensions['I'].width = 22
    ws7.column_dimensions['J'].width = 16
    ws7.column_dimensions['K'].width = 18
    ws7.column_dimensions['L'].width = 16
    ws7.column_dimensions['M'].width = 18
    ws7.column_dimensions['N'].width = 22
    ws7.column_dimensions['O'].width = 30

    out_path = r"g:\ΓΕΦΥΡΕΣ\ERGO ZWHS\ERGO_Insurance_Calculator_and_Coverages.xlsx"
    wb.save(out_path)
    print(f"Successfully generated Excel workbook at: {out_path}")

if __name__ == "__main__":
    build_ergo_excel()
